import streamlit as st
import pandas as pd
import plotly.express as px
import io
import re
import os
import tempfile
from fpdf import FPDF  # NOTA: Requiere tener instalado fpdf2 (pip install fpdf2)
from supabase import create_client, Client
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =================================================================
# 🔌 CONEXIÓN SEGURA AL CENTRO DE DATOS (REGLA DE ORO: INTACTO)
# =================================================================
@st.cache_resource
def iniciar_conexion():
    url = st.secrets["SUPABASE_URL"].replace('"', '').replace("'", "").strip()
    key = st.secrets["SUPABASE_KEY"].replace('"', '').replace("'", "").strip()
    return create_client(url, key)

# =================================================================
# 🖨️ MOTOR GENERADOR DE PDF OPTIMIZADO (SOPORTE NATIVO DE TILDES)
# =================================================================
class GeneradorPDF(FPDF):
    def header(self):
        self.set_font('Helvetica', 'B', 15)
        self.set_text_color(13, 27, 42) # Azul Corporativo
        self.cell(0, 10, 'GÉNESIS OMR - BOLETÍN DE RESULTADOS', 0, 1, 'C')
        self.line(10, 22, 200, 22)
        self.ln(5)

def ensamblar_pdf(datos_estudiante, llave_maestra, nombre_prueba):
    pdf = GeneradorPDF()
    pdf.add_page()
    
    # 1. Cabecera del Estudiante (Ahora con tildes y eñes reales)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(40, 8, 'Estudiante:', 0, 0)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 8, str(datos_estudiante['estudiante']), 0, 1)
    
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(40, 8, 'Evaluación:', 0, 0)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 8, str(nombre_prueba), 0, 1)
    
    pdf.set_font('Helvetica', 'B', 11)
    pdf.cell(40, 8, 'Fecha Escaneo:', 0, 0)
    pdf.set_font('Helvetica', '', 11)
    pdf.cell(0, 8, str(datos_estudiante['fecha_formateada']), 0, 1)
    pdf.ln(5)
    
    # 2. Caja de Calificación
    pdf.set_fill_color(230, 240, 255)
    pdf.set_font('Helvetica', 'B', 13)
    nota_texto = f"CALIFICACIÓN DEFINITIVA: {datos_estudiante['puntaje_obtenido']} / {datos_estudiante['puntaje_maximo']} ({datos_estudiante['porcentaje']}%)"
    pdf.cell(0, 12, nota_texto, 1, 1, 'C', fill=True)
    pdf.ln(8)
    
    # 3. Tabla de Desglose
    pdf.set_font('Helvetica', 'B', 10)
    pdf.set_fill_color(13, 27, 42)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(30, 8, 'Pregunta', 1, 0, 'C', fill=True)
    pdf.cell(30, 8, 'Respuesta', 1, 0, 'C', fill=True)
    pdf.cell(30, 8, 'Correcta', 1, 0, 'C', fill=True)
    pdf.cell(100, 8, 'Tema Evaluado', 1, 1, 'C', fill=True)
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 9)
    
    respuestas_alumno = datos_estudiante['respuestas_json']
    temas_a_reforzar = set()
    
    for item in llave_maestra:
        preg = str(item["Pregunta"])
        correcta = str(item["Respuesta Correcta"])
        tema = str(item.get("Tema", "Concepto General"))
        marcada = str(respuestas_alumno.get(item["Pregunta"], "VACÍA"))
        
        if marcada == correcta:
            pdf.set_fill_color(220, 255, 220) # Verde clarito si acertó
        else:
            pdf.set_fill_color(255, 220, 220) # Rojo clarito si falló
            temas_a_reforzar.add(tema)
            
        pdf.cell(30, 8, preg.replace("Pregunta ", "P "), 1, 0, 'C', fill=True)
        pdf.cell(30, 8, marcada, 1, 0, 'C', fill=True)
        pdf.cell(30, 8, correcta, 1, 0, 'C', fill=True)
        pdf.cell(100, 8, tema, 1, 1, 'L', fill=True)
        
    pdf.ln(8)
    
    # 4. Conclusión y Recomendaciones
    pdf.set_font('Helvetica', 'B', 11)
    if temas_a_reforzar:
        pdf.cell(0, 8, 'PLAN DE MEJORA ACADÉMICA:', 0, 1)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(0, 6, 'El estudiante requiere reforzar urgentemente los siguientes componentes:', 0, 1)
        for t in temas_a_reforzar:
            pdf.cell(5, 6, '-', 0, 0)
            pdf.cell(0, 6, str(t), 0, 1)
    else:
        pdf.cell(0, 8, 'RESULTADO EXCELENTE:', 0, 1)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(
